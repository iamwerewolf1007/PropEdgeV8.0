"""
PropEdge V8 — export_to_excel.py
Exports ALL data from today.json + xlsx database into one flat Excel sheet.
Every play, every field, every signal, every database stat — nothing left behind.

Output: exports/PropEdge_V8_Export_YYYY-MM-DD_HH-MM.xlsx
Creates a new timestamped file every run. Never overwrites old exports.

Usage:
    python3 scripts/export_to_excel.py
"""
import os, sys, json
from datetime import datetime
import pandas as pd
import numpy as np

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT  = os.path.join(SCRIPT_DIR, '..')
TODAY_JSON = os.path.join(REPO_ROOT, 'today.json')
EXPORTS_DIR = os.path.join(REPO_ROOT, 'exports')
NBA_XLSX   = os.path.join(REPO_ROOT, 'source-files', 'NBA_2025_26_Season_Player_Stats.xlsx')

sys.path.insert(0, SCRIPT_DIR)
from model import TEAM_DEF_RANK, _def_tier

# Position-specific DVP (same table as dashboard)
DVP_RAW = {
    'BOS':{'PG':4,'SG':1,'SF':1,'PF':3,'C':1},  'DET':{'PG':1,'SG':7,'SF':10,'PF':17,'C':5},
    'GSW':{'PG':16,'SG':16,'SF':16,'PF':11,'C':21},'ATL':{'PG':12,'SG':27,'SF':27,'PF':22,'C':17},
    'HOU':{'PG':7,'SG':3,'SF':8,'PF':1,'C':7},   'BKN':{'PG':9,'SG':26,'SF':17,'PF':23,'C':19},
    'MEM':{'PG':24,'SG':20,'SF':21,'PF':25,'C':23},'LAC':{'PG':19,'SG':21,'SF':6,'PF':2,'C':12},
    'DAL':{'PG':13,'SG':29,'SF':12,'PF':26,'C':27},'CLE':{'PG':15,'SG':10,'SF':15,'PF':16,'C':20},
    'CHA':{'PG':5,'SG':13,'SF':2,'PF':14,'C':4},  'DEN':{'PG':14,'SG':8,'SF':20,'PF':19,'C':10},
    'IND':{'PG':29,'SG':14,'SF':28,'PF':13,'C':24},'LAL':{'PG':11,'SG':15,'SF':22,'PF':8,'C':6},
    'MIA':{'PG':27,'SG':18,'SF':19,'PF':27,'C':14},'CHI':{'PG':20,'SG':17,'SF':29,'PF':28,'C':26},
    'NOP':{'PG':21,'SG':25,'SF':23,'PF':20,'C':25},'UTA':{'PG':30,'SG':30,'SF':24,'PF':30,'C':22},
    'SAC':{'PG':22,'SG':28,'SF':14,'PF':18,'C':29},'POR':{'PG':18,'SG':22,'SF':26,'PF':15,'C':28},
    'WAS':{'PG':26,'SG':23,'SF':30,'PF':29,'C':30},'OKC':{'PG':2,'SG':11,'SF':13,'PF':4,'C':8},
    'NYK':{'PG':3,'SG':6,'SF':9,'PF':7,'C':2},   'PHI':{'PG':8,'SG':24,'SF':18,'PF':24,'C':15},
    'PHX':{'PG':6,'SG':2,'SF':7,'PF':9,'C':16},  'MIN':{'PG':25,'SG':4,'SF':4,'PF':10,'C':13},
    'ORL':{'PG':23,'SG':12,'SF':3,'PF':12,'C':11},'TOR':{'PG':17,'SG':9,'SF':5,'PF':6,'C':9},
    'SAS':{'PG':10,'SG':5,'SF':11,'PF':5,'C':18}, 'MIL':{'PG':28,'SG':19,'SF':25,'PF':21,'C':3},
}

# ── SIGNAL LABELS ────────────────────────────────────────────────────────────
SIG_LABELS = {
    '1': 'S1_L5_avg_vs_line',
    '2': 'S2_L10_avg_vs_line',
    '3': 'S3_L20_avg_vs_line',
    '4': 'S4_L30_avg_vs_line',
    '5': 'S5_HomeAway_avg_vs_line',
    '6': 'S6_B2B_Rest_avg_vs_line',
    '7': 'S7_DVP_opp_quality',
    '8': 'S8_H2H_vs_opponent',
    '9': 'S9_FG_shooting_trend',
    '10': 'S10_Minutes_trend',
}

def pos_slot(pos):
    p = (pos or '').upper().replace('-','')
    if p in ('PG','G'):    return 'PG'
    if p == 'SG':          return 'SG'
    if p in ('SF','F'):    return 'SF'
    if p == 'PF':          return 'PF'
    if p == 'C':           return 'C'
    if 'GF' in p or 'FG' in p: return 'SG'
    if 'FC' in p or 'CF' in p: return 'C'
    return 'SF'

def dvp_pos_rank(opp, pos):
    d = DVP_RAW.get(opp, {})
    return d.get(pos_slot(pos), 15)

def reasoning(p):
    """Human-readable prematch reasoning summary."""
    sigs = p.get('signals', {})
    over_sigs  = sum(1 for s in sigs.values() if s > 0.6)
    under_sigs = sum(1 for s in sigs.values() if s < 0.4)
    neutral    = 10 - over_sigs - under_sigs
    l10 = p.get('l10'); l30 = p.get('l30'); line = p.get('line',0)
    trend = ''
    if l10 and l30:
        diff = round(l10-l30, 1)
        if diff > 1:   trend = f'Hot (L10 {diff:+.1f} vs L30)'
        elif diff < -1:trend = f'Cold (L10 {diff:+.1f} vs L30)'
        else:          trend = 'Stable form'
    avg_context = ''
    if l10: avg_context = f'L10={l10:.1f}'
    if l30: avg_context += f' L30={l30:.1f}'
    return (f"{p.get('direction','?')} {line} | "
            f"{over_sigs}/10 signals OVER, {under_sigs}/10 UNDER, {neutral} neutral | "
            f"{avg_context} | {trend} | "
            f"DVP #{dvp_pos_rank(p.get('opponent',''), p.get('position',''))} vs {p.get('opponent','')}")

def grade_explanation(p):
    """Human-readable grading explanation."""
    r = p.get('result')
    if not r: return 'Pre-Match — not yet graded'
    actual = p.get('actual_pts')
    line   = p.get('line', 0)
    dirn   = p.get('direction','?')
    if r == 'DNP':  return f'DNP — player did not play'
    if r == 'PUSH': return f'PUSH — scored {actual} = line {line}'
    if actual is None: return r
    diff = round(actual - line, 1)
    hit  = (dirn=='OVER' and actual>line) or (dirn=='UNDER' and actual<line)
    return (f"{r} | Bet {dirn} {line} | Scored {actual} ({diff:+.1f}) | "
            f"{'✓ Correct' if hit else '✗ Wrong direction'}")

def load_db_lookups():
    """Load all xlsx database sheets into fast lookup dicts."""
    xl  = pd.ExcelFile(NBA_XLSX)
    avg = xl.parse('Player Averages')
    ha  = xl.parse('Home Away Splits')
    b2b = xl.parse('B2B vs Rest Splits')
    oq  = xl.parse('Opp Quality Splits')
    h2h = xl.parse('H2H Breakdown')
    sh  = xl.parse('Shooting Trends')
    mn  = xl.parse('Minutes Trends')

    avg_d = {r['Player']: r.to_dict() for _, r in avg.iterrows()}
    ha_d  = {r['Player']: r.to_dict() for _, r in ha.iterrows()}
    b2b_d = {r['Player']: r.to_dict() for _, r in b2b.iterrows()}
    oq_d  = {r['Player']: r.to_dict() for _, r in oq.iterrows()}
    h2h_d = {(r['Player'],r['Opponent']): r.to_dict() for _, r in h2h.iterrows()}
    sh_d  = {r['Player']: r.to_dict() for _, r in sh.iterrows()}
    mn_d  = {r['Player']: r.to_dict() for _, r in mn.iterrows()}
    return avg_d, ha_d, b2b_d, oq_d, h2h_d, sh_d, mn_d

def build_row(p, avg_d, ha_d, b2b_d, oq_d, h2h_d, sh_d, mn_d):
    """Build one flat export row with every data point."""
    player = p.get('player','')
    opp    = p.get('opponent','')
    pos    = p.get('position','')
    line   = p.get('line', 0)
    dirn   = p.get('direction','')
    sigs   = p.get('signals', {})

    # Database rows
    a  = avg_d.get(player, {})
    h  = ha_d.get(player, {})
    b  = b2b_d.get(player, {})
    o  = oq_d.get(player, {})
    hh = h2h_d.get((player, opp), {})
    s  = sh_d.get(player, {})
    m  = mn_d.get(player, {})

    # DVP position rank
    dvp_rank = dvp_pos_rank(opp, pos)

    # Line movement
    hist = p.get('line_history', [line])
    line_moved = len(hist) > 1
    line_move_str = ' → '.join(str(x) for x in hist) if line_moved else ''

    # Recent 20 scores as string
    r20 = p.get('recent20', [])
    r20_str = ','.join(str(x) for x in r20) if r20 else ''
    over_last20  = sum(1 for x in r20 if x > line) if r20 else None
    under_last20 = sum(1 for x in r20 if x < line) if r20 else None

    row = {
        # ── IDENTITY ────────────────────────────────────────────────────
        'Date':                p.get('date',''),
        'Game_Time_ET':        p.get('game_time',''),
        'Player':              player,
        'Team':                p.get('team',''),
        'Position':            pos,
        'Opponent':            opp,
        'Game':                p.get('game',''),
        'Home':                p.get('home',''),
        'Away':                p.get('away',''),
        'Is_Home':             p.get('is_home', False),
        'Is_B2B':              p.get('is_b2b', False),

        # ── PROP LINE ────────────────────────────────────────────────────
        'Line':                line,
        'Min_Line':            p.get('min_line', line),
        'Max_Line':            p.get('max_line', line),
        'Line_History':        line_move_str,
        'Line_Moved':          line_moved,
        'Over_Odds':           p.get('over_odds', -110),
        'Under_Odds':          p.get('under_odds', -110),
        'Books':               p.get('books', 0),
        'Event_ID':            p.get('event_id',''),

        # ── MODEL OUTPUT ─────────────────────────────────────────────────
        'Direction':           dirn,
        'Confidence_Pct':      round(p.get('confidence',0)*100, 2),
        'Prob_Over_Pct':       round(p.get('prob_over',0)*100, 2),
        'Tier':                p.get('tier',''),

        # ── SIGNALS (raw 0–1 sigmoid values) ────────────────────────────
        'S1_L5_vs_Line':       round(float(sigs.get('1',0.5)),4),
        'S2_L10_vs_Line':      round(float(sigs.get('2',0.5)),4),
        'S3_L20_vs_Line':      round(float(sigs.get('3',0.5)),4),
        'S4_L30_vs_Line':      round(float(sigs.get('4',0.5)),4),
        'S5_HomeAway':         round(float(sigs.get('5',0.5)),4),
        'S6_B2B_Rest':         round(float(sigs.get('6',0.5)),4),
        'S7_DVP_Opp_Quality':  round(float(sigs.get('7',0.5)),4),
        'S8_H2H':              round(float(sigs.get('8',0.5)),4),
        'S9_FG_Trend':         round(float(sigs.get('9',0.5)),4),
        'S10_Min_Trend':       round(float(sigs.get('10',0.5)),4),
        'Signals_Over_0.6':    sum(1 for v in sigs.values() if float(v)>0.6),
        'Signals_Under_0.4':   sum(1 for v in sigs.values() if float(v)<0.4),
        'Signals_Neutral':     sum(1 for v in sigs.values() if 0.4<=float(v)<=0.6),

        # ── SCORING AVERAGES (from play dict) ───────────────────────────
        'L5_Avg_PTS':          p.get('l10'),   # note: l5 not stored separately, use DB below
        'L10_Avg_PTS':         p.get('l10'),
        'L20_Avg_PTS':         p.get('l20'),
        'L30_Avg_PTS':         p.get('l30'),

        # ── DATABASE: PLAYER AVERAGES ───────────────────────────────────
        'DB_L3_Avg_PTS':       a.get('L3 Avg PTS'),
        'DB_L5_Avg_PTS':       a.get('L5 Avg PTS'),
        'DB_L10_Avg_PTS':      a.get('L10 Avg PTS'),
        'DB_L20_Avg_PTS':      a.get('L20 Avg PTS'),
        'DB_L30_Avg_PTS':      a.get('L30 Avg PTS'),
        'DB_L50_Avg_PTS':      a.get('L50 Avg PTS'),
        'DB_L5_Avg_MIN':       a.get('L5 Avg MIN'),
        'DB_L10_Avg_MIN':      a.get('L10 Avg MIN'),
        'DB_L30_Avg_MIN':      a.get('L30 Avg MIN'),
        'DB_L10_FG_Pct':       a.get('L10 FG%'),
        'DB_L30_FG_Pct':       a.get('L30 FG%'),
        'DB_Total_Games':      a.get('Total Games'),

        # ── DATABASE: HOME / AWAY SPLITS ────────────────────────────────
        'DB_Home_L10_Avg_PTS': h.get('Home L10 Avg PTS'),
        'DB_Home_All_Avg_PTS': h.get('Home All Avg PTS'),
        'DB_Home_All_Games':   h.get('Home All Games'),
        'DB_Away_L10_Avg_PTS': h.get('Away L10 Avg PTS'),
        'DB_Away_All_Avg_PTS': h.get('Away All Avg PTS'),
        'DB_Away_All_Games':   h.get('Away All Games'),

        # ── DATABASE: B2B / REST SPLITS ─────────────────────────────────
        'DB_B2B_Games':        b.get('B2B Games'),
        'DB_B2B_Avg_PTS':      b.get('B2B Avg PTS'),
        'DB_Rest_Games':       b.get('Rest Games'),
        'DB_Rest_Avg_PTS':     b.get('Rest Avg PTS'),
        'DB_PTS_Diff_Rest_B2B':b.get('PTS Diff (Rest-B2B)'),

        # ── DATABASE: OPP QUALITY ───────────────────────────────────────
        'DB_vs_Top10_Games':   o.get('vs Top-10 Def Games'),
        'DB_vs_Top10_Avg_PTS': o.get('vs Top-10 Def Avg PTS'),
        'DB_vs_Mid_Games':     o.get('vs Mid Def Games'),
        'DB_vs_Mid_Avg_PTS':   o.get('vs Mid Def Avg PTS'),
        'DB_vs_Bot10_Games':   o.get('vs Bot-10 Def Games'),
        'DB_vs_Bot10_Avg_PTS': o.get('vs Bot-10 Def Avg PTS'),

        # ── DATABASE: H2H vs OPPONENT ───────────────────────────────────
        'DB_H2H_Games':        hh.get('H2H Games'),
        'DB_H2H_Avg_PTS':      hh.get('H2H Avg PTS'),
        'DB_H2H_Avg_MIN':      hh.get('H2H Avg MIN'),
        'DB_L5_H2H_Avg_PTS':   hh.get('L5 H2H Avg PTS'),
        'DB_L3_H2H_Avg_PTS':   hh.get('L3 H2H Avg PTS'),
        'DB_H2H_FG_Pct':       hh.get('H2H FG%'),
        'DB_H2H_PTS_Trend_L3_vs_All': hh.get('H2H PTS Trend (L3 vs All)'),

        # ── DATABASE: SHOOTING TRENDS ───────────────────────────────────
        'DB_FG_Trend_L10_L30': s.get('FG% Trend (L10-L30)'),
        'DB_3P_Trend_L10_L30': s.get('3P% Trend (L10-L30)'),
        'DB_L10_FGA':          s.get('L10 Avg FGA'),
        'DB_L30_FGA':          s.get('L30 Avg FGA'),

        # ── DATABASE: MINUTES TRENDS ────────────────────────────────────
        'DB_Min_Trend_L10_L30':m.get('MIN Trend (L10-L30)'),
        'DB_L10_Avg_MIN':      m.get('L10 Avg MIN'),
        'DB_L30_Avg_MIN':      m.get('L30 Avg MIN'),
        'DB_Likely_Starter':   m.get('Likely Starter'),

        # ── DVP ─────────────────────────────────────────────────────────
        'DVP_Opp_Position_Rank':    dvp_rank,
        'DVP_Tier':            _def_tier(opp),

        # ── RECENT GAMES ────────────────────────────────────────────────
        'Recent20_Scores':     r20_str,
        'Recent20_Over_Line':  over_last20,
        'Recent20_Under_Line': under_last20,
        'Recent20_Hit_Rate_Pct': round(over_last20/len(r20)*100, 1) if r20 and dirn=='OVER' and over_last20 is not None
                               else (round(under_last20/len(r20)*100,1) if r20 and dirn=='UNDER' and under_last20 is not None else None),

        # ── PREMATCH REASONING ──────────────────────────────────────────
        'Prematch_Reasoning':  reasoning(p),

        # ── GRADING ─────────────────────────────────────────────────────
        'Result':              p.get('result','Pre-Match'),
        'Actual_PTS':          p.get('actual_pts'),
        'Points_vs_Line':      round(float(p['actual_pts']) - line, 1) if p.get('actual_pts') is not None else None,
        'Grade_Explanation':   grade_explanation(p),
        'Is_Correct':          (
            (dirn=='OVER'  and p.get('actual_pts') is not None and float(p['actual_pts']) > line) or
            (dirn=='UNDER' and p.get('actual_pts') is not None and float(p['actual_pts']) < line)
        ) if p.get('result') in ('WIN','LOSS') else None,
    }
    return row


if __name__ == '__main__':
    print("="*60)
    print("PropEdge V8 — export_to_excel.py")
    print("="*60)

    # ── Load data ────────────────────────────────────────────────
    print("\n[1/4] Loading today.json...")
    if not os.path.exists(TODAY_JSON):
        print("  today.json not found. Run run_everything.py first.")
        sys.exit(1)
    with open(TODAY_JSON) as f:
        plays = json.load(f)
    print(f"  {len(plays)} plays loaded")

    print("\n[2/4] Loading database lookups...")
    avg_d, ha_d, b2b_d, oq_d, h2h_d, sh_d, mn_d = load_db_lookups()

    print("\n[3/4] Building export rows...")
    rows = []
    for p in plays:
        try:
            rows.append(build_row(p, avg_d, ha_d, b2b_d, oq_d, h2h_d, sh_d, mn_d))
        except Exception as e:
            print(f"  WARNING: {p.get('player','')} {p.get('date','')} — {e}")
    print(f"  Built {len(rows)} rows")

    df = pd.DataFrame(rows)

    # Sort by date desc, then tier, then confidence desc
    tier_ord = {'T1':0,'T2':1,'T3':2}
    df['_tier_ord'] = df['Tier'].map(tier_ord).fillna(3)
    df = df.sort_values(['Date','_tier_ord','Confidence_Pct'],
                        ascending=[False, True, False]).drop(columns=['_tier_ord'])

    print("\n[4/4] Writing Excel...")
    os.makedirs(EXPORTS_DIR, exist_ok=True)
    ts   = datetime.now().strftime('%Y-%m-%d_%H-%M')
    path = os.path.join(EXPORTS_DIR, f'PropEdge_V8_Export_{ts}.xlsx')

    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='All Data', index=False)
        ws = writer.sheets['All Data']

        # ── Freeze header + auto-width ────────────────────────────────
        ws.freeze_panes = 'A2'
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

        # ── Colour code Result column ─────────────────────────────────
        from openpyxl.styles import PatternFill, Font, Alignment
        res_col = df.columns.get_loc('Result') + 1  # 1-indexed
        tier_col = df.columns.get_loc('Tier') + 1
        fills = {
            'WIN':      PatternFill('solid', fgColor='00D6A4'),
            'LOSS':     PatternFill('solid', fgColor='FF6B6B'),
            'PUSH':     PatternFill('solid', fgColor='AAAAAA'),
            'DNP':      PatternFill('solid', fgColor='DDDDDD'),
            'T1':       PatternFill('solid', fgColor='003D2B'),
            'T2':       PatternFill('solid', fgColor='001D40'),
            'T3':       PatternFill('solid', fgColor='1A1A1A'),
        }
        for row_idx in range(2, len(df)+2):
            res_cell  = ws.cell(row=row_idx, column=res_col)
            tier_cell = ws.cell(row=row_idx, column=tier_col)
            res_val   = str(res_cell.value or '')
            tier_val  = str(tier_cell.value or '')
            if res_val in fills:
                res_cell.fill = fills[res_val]
                res_cell.font = Font(bold=True, color='FFFFFF')
            if tier_val in fills:
                tier_cell.fill = fills[tier_val]
                tier_cell.font = Font(bold=True,
                    color='00E5A0' if tier_val=='T1' else '4DA6FF' if tier_val=='T2' else '888888')

        # ── Bold + centre header row ───────────────────────────────────
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

    print(f"\n  ✓ Saved: {path}")
    print(f"  Rows: {len(df)} | Columns: {len(df.columns)}")

    graded = df[df['Result'].isin(['WIN','LOSS'])]
    wins   = (graded['Result']=='WIN').sum()
    print(f"\n  Season summary in export:")
    print(f"  Graded: {len(graded)} | HR: {round(wins/len(graded)*100,1)}%" if len(graded) else "  No graded plays")
    for tier in ['T1','T2','T3']:
        tg = graded[graded['Tier']==tier]
        tw = (tg['Result']=='WIN').sum()
        print(f"  {tier}: {len(tg)} plays | {round(tw/len(tg)*100,1)}% HR" if len(tg) else f"  {tier}: 0 plays")
    print("\nDone.")
