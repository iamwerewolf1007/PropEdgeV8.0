"""
PropEdge V8 — xlsx Engine
Reads and writes the NBA stats xlsx database.
Single source of truth for all lookup tables.
"""

import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

NBA_XLSX  = os.path.join(os.path.dirname(__file__), '..', 'source-files',
                         'NBA_2025_26_Season_Player_Stats.xlsx')
PROP_XLSX = os.path.join(os.path.dirname(__file__), '..', 'source-files',
                         'PropEdge_-_Match_and_Player_Prop_lines_.xlsx')


# ── READ ──────────────────────────────────────────────────────────────────────

def load_all_sheets() -> dict:
    """
    Load all sheets from NBA xlsx. Returns dict of DataFrames.
    Keys: 'gl', 'avg', 'ha', 'wl', 'b2b', 'oq', 'h2h', 'shoot', 'mins'
    """
    xl  = pd.ExcelFile(NBA_XLSX)
    gl  = xl.parse('All Game Logs')
    gl['Date'] = pd.to_datetime(gl['Date'])

    return {
        'gl':    gl,
        'avg':   xl.parse('Player Averages'),
        'ha':    xl.parse('Home Away Splits'),
        'wl':    xl.parse('Win Loss Splits'),
        'b2b':   xl.parse('B2B vs Rest Splits'),
        'oq':    xl.parse('Opp Quality Splits'),
        'h2h':   xl.parse('H2H Breakdown'),
        'shoot': xl.parse('Shooting Trends'),
        'mins':  xl.parse('Minutes Trends'),
    }


def load_props() -> tuple:
    """Returns (props_df, spreads_df) from the props xlsx."""
    xl      = pd.ExcelFile(PROP_XLSX)
    props   = xl.parse('Player_Points_Props')
    spreads = xl.parse('Team_Spreads_Totals')
    props['Date']   = pd.to_datetime(props['Date'])
    spreads['Date'] = pd.to_datetime(spreads['Date'])
    return props, spreads


def get_player_rows(sheets: dict, player: str) -> dict:
    """
    Retrieve all per-player rows needed for scoring.
    Returns dict with keys: avg, ha, b2b, oq, shoot, mins
    h2h is keyed separately per opponent via get_h2h_row().
    """
    def _first(df, key_col='Player'):
        rows = df[df[key_col] == player]
        return rows.iloc[0].to_dict() if len(rows) > 0 else {}

    return {
        'avg':   _first(sheets['avg']),
        'ha':    _first(sheets['ha']),
        'b2b':   _first(sheets['b2b']),
        'oq':    _first(sheets['oq']),
        'shoot': _first(sheets['shoot']),
        'mins':  _first(sheets['mins']),
    }


def get_h2h_row(sheets: dict, player: str, opponent: str) -> dict:
    """Return H2H row for player vs opponent, or empty dict."""
    h2h = sheets['h2h']
    rows = h2h[(h2h['Player'] == player) & (h2h['Opponent'] == opponent)]
    return rows.iloc[0].to_dict() if len(rows) > 0 else {}


def get_player_team(sheets: dict, player: str) -> str:
    """Look up player's current team from Player Averages sheet."""
    avg = sheets['avg']
    rows = avg[avg['Player'] == player]
    if len(rows) > 0:
        return str(rows.iloc[0]['Team'])
    # fallback: check game logs for most recent team
    gl = sheets['gl']
    p_logs = gl[gl['Player'] == player].sort_values('Date', ascending=False)
    if len(p_logs) > 0:
        return str(p_logs.iloc[0]['Team'])
    return ''


# ── WRITE ─────────────────────────────────────────────────────────────────────

def _overwrite_sheet(wb, sheet_name: str, df: pd.DataFrame):
    """Overwrite an xlsx sheet with new DataFrame content."""
    ws = wb[sheet_name]
    ws.delete_rows(1, ws.max_row + 1)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, val in enumerate(row, 1):
            # Convert numpy types to native Python for openpyxl
            if isinstance(val, (np.integer,)):
                val = int(val)
            elif isinstance(val, (np.floating,)):
                val = float(val)
            elif isinstance(val, (np.bool_,)):
                val = bool(val)
            ws.cell(row=r_idx, column=c_idx, value=val)


def save_nba_sheets(sheets: dict):
    """Write all 9 NBA sheets back to the xlsx file."""
    wb = load_workbook(NBA_XLSX)

    sheet_map = {
        'gl':    'All Game Logs',
        'avg':   'Player Averages',
        'ha':    'Home Away Splits',
        'wl':    'Win Loss Splits',
        'b2b':   'B2B vs Rest Splits',
        'oq':    'Opp Quality Splits',
        'h2h':   'H2H Breakdown',
        'shoot': 'Shooting Trends',
        'mins':  'Minutes Trends',
    }

    for key, sheet_name in sheet_map.items():
        if key in sheets:
            _overwrite_sheet(wb, sheet_name, sheets[key])

    wb.save(NBA_XLSX)
    wb.close()


def append_game_logs(new_rows: pd.DataFrame):
    """
    Append new game log rows to All Game Logs sheet.
    Also triggers full recomputation of all derived sheets.
    Deduplicates by Player + Date before appending.
    """
    sheets = load_all_sheets()
    gl     = sheets['gl']

    # Dedup: remove any existing rows for the same Player+Date combos
    new_rows['Date'] = pd.to_datetime(new_rows['Date'])
    existing_keys = set(zip(gl['Player'], gl['Date'].dt.date))
    mask = new_rows.apply(
        lambda r: (r['Player'], pd.to_datetime(r['Date']).date()) not in existing_keys,
        axis=1
    )
    to_add = new_rows[mask].copy()

    if len(to_add) == 0:
        print("No new rows to append (all already exist)")
        return

    gl_new = pd.concat([gl, to_add], ignore_index=True)
    gl_new = gl_new.sort_values(['Player', 'Date']).reset_index(drop=True)
    sheets['gl'] = gl_new

    print(f"Appended {len(to_add)} new game log rows")

    # Recompute all derived sheets from clean game logs
    recompute_all_sheets(sheets)
    save_nba_sheets(sheets)
    print("All sheets recomputed and saved")


def recompute_all_sheets(sheets: dict):
    """
    Recompute all 8 derived sheets from game logs.
    Updates sheets dict in-place.
    Only uses rows where Minutes > 0 for stat averages.
    """
    gl = sheets['gl']
    gl_played = gl[gl['Minutes'] > 0].copy()

    sheets['avg']   = _compute_player_averages(gl_played)
    sheets['ha']    = _compute_home_away(gl_played)
    sheets['wl']    = _compute_win_loss(gl_played)
    sheets['b2b']   = _compute_b2b(gl_played)
    sheets['oq']    = _compute_opp_quality(gl_played)
    sheets['h2h']   = _compute_h2h(gl_played)
    sheets['shoot'] = _compute_shooting(gl_played)
    sheets['mins']  = _compute_minutes(gl_played)


# ── SHEET COMPUTATION FUNCTIONS ───────────────────────────────────────────────

def _safe_mean(s):
    return round(float(s.mean()), 2) if len(s) > 0 else np.nan

def _safe_pct(made, att):
    total = att.sum()
    return round(float(made.sum() / total) * 100, 2) if total > 0 else np.nan

def _safe_pct_dec(made, att):
    total = att.sum()
    return round(float(made.sum() / total), 4) if total > 0 else np.nan


def _compute_player_averages(gl: pd.DataFrame) -> pd.DataFrame:
    """Compute Player Averages sheet from clean game logs."""
    windows = [200, 100, 50, 30, 20, 10, 5, 3]
    gl_s = gl.sort_values(['Player', 'Date'], ascending=[True, False])
    rows = []
    for player, grp in gl_s.groupby('Player'):
        row = {
            'Player': player,
            'Team':   grp.iloc[0]['Team'],
            'Total Games': len(grp),
        }
        for w in windows:
            sub = grp.head(w)
            row[f'L{w} Avg PTS'] = _safe_mean(sub['Points'])
            row[f'L{w} Avg MIN'] = _safe_mean(sub['Minutes'])
            row[f'L{w} Avg FGA'] = _safe_mean(sub['FGA'])
            row[f'L{w} FG%']     = _safe_pct(sub['FGM'], sub['FGA'])
            row[f'L{w} 3P%']     = _safe_pct(sub['3PM'], sub['3PA'])
            row[f'L{w} FT%']     = _safe_pct(sub['FTM'], sub['FTA'])
            row[f'L{w} Avg REB'] = _safe_mean(sub['REB'])
            row[f'L{w} Avg AST'] = _safe_mean(sub['AST'])
            row[f'L{w} Avg FTA'] = _safe_mean(sub['FTA'])
        rows.append(row)
    return pd.DataFrame(rows)


def _compute_home_away(gl: pd.DataFrame) -> pd.DataFrame:
    """Compute Home Away Splits sheet."""
    gl_s = gl.sort_values(['Player', 'Date'], ascending=[True, False])
    rows = []
    for player, grp in gl_s.groupby('Player'):
        home = grp[grp['Home/Away'] == 'Home']
        away = grp[grp['Home/Away'] == 'Away']
        row = {'Player': player, 'Team': grp.iloc[0]['Team']}
        for label, sub in [('Home', home), ('Away', away)]:
            for win, w_lbl in [('All', None), ('L30', 30), ('L10', 10), ('L5', 5)]:
                s = sub.head(w_lbl) if w_lbl else sub
                row[f'{label} {win} Games']   = len(s)
                row[f'{label} {win} Avg PTS'] = _safe_mean(s['Points']) if len(s) > 0 else np.nan
                row[f'{label} {win} Avg MIN'] = _safe_mean(s['Minutes']) if len(s) > 0 else np.nan
                row[f'{label} {win} FG%']     = _safe_pct(s['FGM'], s['FGA']) if len(s) > 0 else np.nan
                row[f'{label} {win} 3P%']     = _safe_pct(s['3PM'], s['3PA']) if len(s) > 0 else np.nan
        rows.append(row)
    return pd.DataFrame(rows)


def _compute_win_loss(gl: pd.DataFrame) -> pd.DataFrame:
    """Compute Win Loss Splits sheet."""
    rows = []
    for player, grp in gl.groupby('Player'):
        wins   = grp[grp['W/L'] == 'W']
        losses = grp[grp['W/L'] == 'L']
        row = {
            'Player':    player,
            'Team':      grp.iloc[0]['Team'],
            'Total W':   len(wins),
            'Total L':   len(losses),
            'Win Avg PTS':  _safe_mean(wins['Points'])   if len(wins) > 0 else np.nan,
            'Win Avg MIN':  _safe_mean(wins['Minutes'])  if len(wins) > 0 else np.nan,
            'Win FG%':      _safe_pct(wins['FGM'], wins['FGA']) if len(wins) > 0 else np.nan,
            'Loss Avg PTS': _safe_mean(losses['Points'])  if len(losses) > 0 else np.nan,
            'Loss Avg MIN': _safe_mean(losses['Minutes']) if len(losses) > 0 else np.nan,
            'Loss FG%':     _safe_pct(losses['FGM'], losses['FGA']) if len(losses) > 0 else np.nan,
        }
        w_avg = row['Win Avg PTS']  if pd.notna(row.get('Win Avg PTS',  np.nan)) else None
        l_avg = row['Loss Avg PTS'] if pd.notna(row.get('Loss Avg PTS', np.nan)) else None
        row['PTS Diff (W-L)'] = round(w_avg - l_avg, 2) if (w_avg and l_avg) else np.nan
        rows.append(row)
    return pd.DataFrame(rows)


def _compute_b2b(gl: pd.DataFrame) -> pd.DataFrame:
    """Compute B2B vs Rest Splits sheet."""
    rows = []
    for player, grp in gl.groupby('Player'):
        b2b  = grp[grp['B2B'] == True]
        rest = grp[grp['B2B'] == False]
        b2b_avg  = _safe_mean(b2b['Points'])   if len(b2b) > 0 else np.nan
        rest_avg = _safe_mean(rest['Points'])   if len(rest) > 0 else np.nan
        row = {
            'Player':           player,
            'Team':             grp.iloc[0]['Team'],
            'B2B Games':        len(b2b),
            'B2B Avg PTS':      b2b_avg,
            'B2B Avg MIN':      _safe_mean(b2b['Minutes'])  if len(b2b) > 0 else np.nan,
            'B2B FG%':          _safe_pct(b2b['FGM'], b2b['FGA']) if len(b2b) > 0 else np.nan,
            'Rest Games':       len(rest),
            'Rest Avg PTS':     rest_avg,
            'Rest Avg MIN':     _safe_mean(rest['Minutes']) if len(rest) > 0 else np.nan,
            'Rest FG%':         _safe_pct(rest['FGM'], rest['FGA']) if len(rest) > 0 else np.nan,
            'PTS Diff (Rest-B2B)': round(rest_avg - b2b_avg, 2) if (pd.notna(rest_avg) and pd.notna(b2b_avg)) else np.nan,
        }
        rows.append(row)
    return pd.DataFrame(rows)


def _compute_opp_quality(gl: pd.DataFrame) -> pd.DataFrame:
    """Compute Opp Quality Splits (Top-10/Mid/Bot-10 defense) sheet."""
    rows = []
    for player, grp in gl.groupby('Player'):
        top  = grp[grp['Opp Def Rank'] <= 10]
        bot  = grp[grp['Opp Def Rank'] >= 21]
        mid  = grp[(grp['Opp Def Rank'] > 10) & (grp['Opp Def Rank'] < 21)]
        t_avg = _safe_mean(top['Points'])  if len(top) > 0 else np.nan
        b_avg = _safe_mean(bot['Points'])  if len(bot) > 0 else np.nan
        row = {
            'Player':                  player,
            'Team':                    grp.iloc[0]['Team'],
            'vs Top-10 Def Games':     len(top),
            'vs Top-10 Def Avg PTS':   t_avg,
            'vs Top-10 Def FG%':       _safe_pct(top['FGM'], top['FGA']) if len(top) > 0 else np.nan,
            'vs Mid Def Games':        len(mid),
            'vs Mid Def Avg PTS':      _safe_mean(mid['Points'])  if len(mid) > 0 else np.nan,
            'vs Bot-10 Def Games':     len(bot),
            'vs Bot-10 Def Avg PTS':   b_avg,
            'vs Bot-10 Def FG%':       _safe_pct(bot['FGM'], bot['FGA']) if len(bot) > 0 else np.nan,
            'PTS Diff (Weak-Tough)':   round(b_avg - t_avg, 2) if (pd.notna(b_avg) and pd.notna(t_avg)) else np.nan,
        }
        rows.append(row)
    return pd.DataFrame(rows)


def _compute_h2h(gl: pd.DataFrame) -> pd.DataFrame:
    """Compute H2H Breakdown sheet (player vs each opponent they've faced)."""
    gl_s = gl.sort_values(['Player', 'Opponent', 'Date'], ascending=[True, True, False])
    rows = []
    for (player, opp), grp in gl_s.groupby(['Player', 'Opponent']):
        l5  = grp.head(5)
        l3  = grp.head(3)
        home = grp[grp['Home/Away'] == 'Home']
        away = grp[grp['Home/Away'] == 'Away']
        wins = grp[grp['W/L'] == 'W']
        losses = grp[grp['W/L'] == 'L']
        b2b_h2h  = grp[grp['B2B'] == True]
        rest_h2h = grp[grp['B2B'] == False]

        all_avg  = _safe_mean(grp['Points'])
        l3_avg   = _safe_mean(l3['Points']) if len(l3) > 0 else np.nan

        row = {
            'Player':             player,
            'Team':               grp.iloc[0]['Team'],
            'Opponent':           opp,
            'H2H Games':          len(grp),
            'H2H Avg PTS':        all_avg,
            'H2H Avg MIN':        _safe_mean(grp['Minutes']),
            'H2H Avg FGM':        _safe_mean(grp['FGM']),
            'H2H Avg FGA':        _safe_mean(grp['FGA']),
            'H2H FG%':            _safe_pct(grp['FGM'], grp['FGA']),
            'H2H Avg 3PM':        _safe_mean(grp['3PM']),
            'H2H Avg 3PA':        _safe_mean(grp['3PA']),
            'H2H 3P%':            _safe_pct(grp['3PM'], grp['3PA']),
            'H2H Avg FTM':        _safe_mean(grp['FTM']),
            'H2H Avg FTA':        _safe_mean(grp['FTA']),
            'H2H FT%':            _safe_pct(grp['FTM'], grp['FTA']),
            'H2H Avg REB':        _safe_mean(grp['REB']),
            'H2H Avg AST':        _safe_mean(grp['AST']),
            'H2H Avg STL':        _safe_mean(grp['STL']),
            'H2H Avg BLK':        _safe_mean(grp['BLK']),
            'H2H Avg TOV':        _safe_mean(grp['TOV']),
            'H2H Avg +/-':        _safe_mean(grp['+/-']),
            'L5 H2H Games':       len(l5),
            'L5 H2H Avg PTS':     _safe_mean(l5['Points'])  if len(l5) > 0 else np.nan,
            'L5 H2H Avg MIN':     _safe_mean(l5['Minutes']) if len(l5) > 0 else np.nan,
            'L5 H2H Avg REB':     _safe_mean(l5['REB'])     if len(l5) > 0 else np.nan,
            'L5 H2H Avg AST':     _safe_mean(l5['AST'])     if len(l5) > 0 else np.nan,
            'L5 H2H FG%':         _safe_pct(l5['FGM'], l5['FGA']) if len(l5) > 0 else np.nan,
            'L5 H2H 3P%':         _safe_pct(l5['3PM'], l5['3PA']) if len(l5) > 0 else np.nan,
            'L5 H2H Avg 3PM':     _safe_mean(l5['3PM'])    if len(l5) > 0 else np.nan,
            'L5 H2H Avg 3PA':     _safe_mean(l5['3PA'])    if len(l5) > 0 else np.nan,
            'L5 H2H Avg +/-':     _safe_mean(l5['+/-'])    if len(l5) > 0 else np.nan,
            'L3 H2H Games':       len(l3),
            'L3 H2H Avg PTS':     _safe_mean(l3['Points'])  if len(l3) > 0 else np.nan,
            'L3 H2H Avg MIN':     _safe_mean(l3['Minutes']) if len(l3) > 0 else np.nan,
            'L3 H2H Avg REB':     _safe_mean(l3['REB'])     if len(l3) > 0 else np.nan,
            'L3 H2H Avg AST':     _safe_mean(l3['AST'])     if len(l3) > 0 else np.nan,
            'L3 H2H FG%':         _safe_pct(l3['FGM'], l3['FGA']) if len(l3) > 0 else np.nan,
            'L3 H2H 3P%':         _safe_pct(l3['3PM'], l3['3PA']) if len(l3) > 0 else np.nan,
            'L3 H2H Avg 3PM':     _safe_mean(l3['3PM'])    if len(l3) > 0 else np.nan,
            'L3 H2H Avg 3PA':     _safe_mean(l3['3PA'])    if len(l3) > 0 else np.nan,
            'L3 H2H Avg +/-':     _safe_mean(l3['+/-'])    if len(l3) > 0 else np.nan,
            'H2H Home Games':     len(home),
            'H2H Home Avg PTS':   _safe_mean(home['Points'])  if len(home) > 0 else np.nan,
            'H2H Home Avg MIN':   _safe_mean(home['Minutes']) if len(home) > 0 else np.nan,
            'H2H Home FG%':       _safe_pct(home['FGM'], home['FGA']) if len(home) > 0 else np.nan,
            'H2H Home 3P%':       _safe_pct(home['3PM'], home['3PA']) if len(home) > 0 else np.nan,
            'H2H Home Avg REB':   _safe_mean(home['REB'])  if len(home) > 0 else np.nan,
            'H2H Home Avg AST':   _safe_mean(home['AST'])  if len(home) > 0 else np.nan,
            'H2H Away Games':     len(away),
            'H2H Away Avg PTS':   _safe_mean(away['Points'])  if len(away) > 0 else np.nan,
            'H2H Away Avg MIN':   _safe_mean(away['Minutes']) if len(away) > 0 else np.nan,
            'H2H Away FG%':       _safe_pct(away['FGM'], away['FGA']) if len(away) > 0 else np.nan,
            'H2H Away 3P%':       _safe_pct(away['3PM'], away['3PA']) if len(away) > 0 else np.nan,
            'H2H Away Avg REB':   _safe_mean(away['REB'])  if len(away) > 0 else np.nan,
            'H2H Away Avg AST':   _safe_mean(away['AST'])  if len(away) > 0 else np.nan,
            'H2H Wins':           len(wins),
            'H2H Losses':         len(losses),
            'H2H Win Avg PTS':    _safe_mean(wins['Points'])   if len(wins) > 0 else np.nan,
            'H2H Win FG%':        _safe_pct(wins['FGM'], wins['FGA']) if len(wins) > 0 else np.nan,
            'H2H Loss Avg PTS':   _safe_mean(losses['Points'])  if len(losses) > 0 else np.nan,
            'H2H Loss FG%':       _safe_pct(losses['FGM'], losses['FGA']) if len(losses) > 0 else np.nan,
            'H2H B2B Games':      len(b2b_h2h),
            'H2H B2B Avg PTS':    _safe_mean(b2b_h2h['Points'])  if len(b2b_h2h) > 0 else np.nan,
            'H2H B2B Avg MIN':    _safe_mean(b2b_h2h['Minutes']) if len(b2b_h2h) > 0 else np.nan,
            'H2H Rest Games':     len(rest_h2h),
            'H2H Rest Avg PTS':   _safe_mean(rest_h2h['Points'])  if len(rest_h2h) > 0 else np.nan,
            'H2H Rest Avg MIN':   _safe_mean(rest_h2h['Minutes']) if len(rest_h2h) > 0 else np.nan,
            'H2H PTS Trend (L3 vs All)': round(l3_avg - all_avg, 2) if (pd.notna(l3_avg) and pd.notna(all_avg)) else np.nan,
            'H2H Home-Away PTS Diff': round(
                _safe_mean(home['Points']) - _safe_mean(away['Points']), 2
            ) if (len(home) > 0 and len(away) > 0) else np.nan,
            'H2H W-L PTS Diff': round(
                _safe_mean(wins['Points']) - _safe_mean(losses['Points']), 2
            ) if (len(wins) > 0 and len(losses) > 0) else np.nan,
        }
        rows.append(row)
    return pd.DataFrame(rows)


def _compute_shooting(gl: pd.DataFrame) -> pd.DataFrame:
    """Compute Shooting Trends sheet."""
    gl_s = gl.sort_values(['Player', 'Date'], ascending=[True, False])
    rows = []
    for player, grp in gl_s.groupby('Player'):
        row = {
            'Player':      player,
            'Team':        grp.iloc[0]['Team'],
            'Total Games': len(grp),
        }
        for w in [30, 10, 5, 3]:
            sub = grp.head(w)
            row[f'L{w} FG%']     = _safe_pct(sub['FGM'], sub['FGA'])
            row[f'L{w} 3P%']     = _safe_pct(sub['3PM'], sub['3PA'])
            row[f'L{w} FT%']     = _safe_pct(sub['FTM'], sub['FTA'])
            row[f'L{w} Avg FGA'] = _safe_mean(sub['FGA'])
            row[f'L{w} Avg 3PA'] = _safe_mean(sub['3PA'])
            row[f'L{w} Avg FTA'] = _safe_mean(sub['FTA'])
        # Trends: L10 minus L30
        for stat in ['FG%', '3P%', 'FT%']:
            l10 = row.get(f'L10 {stat}')
            l30 = row.get(f'L30 {stat}')
            row[f'{stat} Trend (L10-L30)'] = round(l10 - l30, 2) if (l10 and l30 and pd.notna(l10) and pd.notna(l30)) else np.nan
        rows.append(row)
    return pd.DataFrame(rows)


def _compute_minutes(gl: pd.DataFrame) -> pd.DataFrame:
    """Compute Minutes Trends sheet."""
    gl_s = gl.sort_values(['Player', 'Date'], ascending=[True, False])
    rows = []
    for player, grp in gl_s.groupby('Player'):
        b2b  = grp[grp['B2B'] == True]
        rest = grp[grp['B2B'] == False]
        l10  = grp.head(10)
        l30  = grp.head(30)
        l10_min = _safe_mean(l10['Minutes']) if len(l10) > 0 else np.nan
        l30_min = _safe_mean(l30['Minutes']) if len(l30) > 0 else np.nan
        row = {
            'Player':           player,
            'Team':             grp.iloc[0]['Team'],
            'Total Games':      len(grp),
            'Season Avg MIN':   _safe_mean(grp['Minutes']),
            'L50 Avg MIN':      _safe_mean(grp.head(50)['Minutes']),
            'L30 Avg MIN':      l30_min,
            'L20 Avg MIN':      _safe_mean(grp.head(20)['Minutes']),
            'L10 Avg MIN':      l10_min,
            'L5 Avg MIN':       _safe_mean(grp.head(5)['Minutes']),
            'L3 Avg MIN':       _safe_mean(grp.head(3)['Minutes']),
            'MIN Trend (L10-L30)': round(l10_min - l30_min, 2) if (pd.notna(l10_min) and pd.notna(l30_min)) else np.nan,
            'Likely Starter':   bool(_safe_mean(grp.head(10)['Minutes']) >= 20) if len(grp) >= 5 else False,
            'B2B Avg MIN':      _safe_mean(b2b['Minutes'])  if len(b2b) > 0 else np.nan,
            'Rest Avg MIN':     _safe_mean(rest['Minutes']) if len(rest) > 0 else np.nan,
        }
        rows.append(row)
    return pd.DataFrame(rows)
